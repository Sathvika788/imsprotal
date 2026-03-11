from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.dynamo import (
    user_repo, log_repo, attendance_repo, task_repo, 
    expense_repo, stipend_repo, project_repo, leave_repo, wfh_repo
)


def generate_daily_report(date_str: str) -> BytesIO:
    """Generate daily report for a specific date"""
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws_summary['A1'] = f'DAILY REPORT - {date_str}'
    ws_summary['A1'].font = Font(bold=True, size=16, color="0066CC")
    ws_summary.merge_cells('A1:E1')
    
    # Get all interns
    interns = user_repo.get_users_by_role('intern')
    
    # Attendance Summary
    ws_summary['A3'] = 'ATTENDANCE SUMMARY'
    ws_summary['A3'].font = header_font
    ws_summary['A3'].fill = header_fill
    
    headers = ['Intern Name', 'Email', 'Status', 'Log Submitted', 'Tasks Completed']
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row = 5
    total_present = 0
    total_late = 0
    total_absent = 0
    
    for intern in interns:
        # Get attendance
        attendance_records = attendance_repo.get_attendance_by_intern(intern['id'])
        attendance = next((a for a in attendance_records if a['date'] == date_str), None)
        status = attendance['status'] if attendance else 'absent'
        
        # Get log
        log = log_repo.get_log_by_intern_and_date(intern['id'], date_str)
        log_status = 'Yes' if log else 'No'
        
        # Get tasks completed today
        tasks = task_repo.get_tasks_by_intern(intern['id'])
        completed_today = len([t for t in tasks if t.get('completed_at', '')[:10] == date_str])
        
        ws_summary.cell(row=row, column=1).value = intern['name']
        ws_summary.cell(row=row, column=2).value = intern['email']
        ws_summary.cell(row=row, column=3).value = status.upper()
        ws_summary.cell(row=row, column=4).value = log_status
        ws_summary.cell(row=row, column=5).value = completed_today
        
        # Color code status
        status_cell = ws_summary.cell(row=row, column=3)
        if status == 'present':
            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            total_present += 1
        elif status == 'late':
            status_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            total_late += 1
        else:
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            total_absent += 1
        
        row += 1
    
    # Summary stats
    ws_summary[f'A{row+2}'] = 'STATISTICS'
    ws_summary[f'A{row+2}'].font = Font(bold=True, size=12)
    
    ws_summary[f'A{row+3}'] = 'Total Interns:'
    ws_summary[f'B{row+3}'] = len(interns)
    ws_summary[f'A{row+4}'] = 'Present:'
    ws_summary[f'B{row+4}'] = total_present
    ws_summary[f'A{row+5}'] = 'Late:'
    ws_summary[f'B{row+5}'] = total_late
    ws_summary[f'A{row+6}'] = 'Absent:'
    ws_summary[f'B{row+6}'] = total_absent
    
    # Adjust column widths
    for col in range(1, 6):
        ws_summary.column_dimensions[get_column_letter(col)].width = 25
    
    # Logs Sheet
    ws_logs = wb.create_sheet("Work Logs")
    ws_logs['A1'] = 'WORK LOGS'
    ws_logs['A1'].font = Font(bold=True, size=14)
    
    log_headers = ['Intern Name', 'Email', 'Status', 'Content', 'Verified By']
    for col, header in enumerate(log_headers, start=1):
        cell = ws_logs.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    log_row = 3
    for intern in interns:
        log = log_repo.get_log_by_intern_and_date(intern['id'], date_str)
        if log:
            ws_logs.cell(row=log_row, column=1).value = intern['name']
            ws_logs.cell(row=log_row, column=2).value = intern['email']
            ws_logs.cell(row=log_row, column=3).value = log.get('status', 'pending').upper()
            ws_logs.cell(row=log_row, column=4).value = log.get('content', '')
            ws_logs.cell(row=log_row, column=5).value = log.get('verified_by', 'N/A')
            log_row += 1
    
    for col in range(1, 6):
        ws_logs.column_dimensions[get_column_letter(col)].width = 30
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_weekly_report(start_date_str: str, end_date_str: str) -> BytesIO:
    """Generate weekly report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = f'WEEKLY REPORT: {start_date_str} to {end_date_str}'
    ws['A1'].font = Font(bold=True, size=16, color="0066CC")
    ws.merge_cells('A1:H1')
    
    # Get all interns
    interns = user_repo.get_users_by_role('intern')
    
    # Headers
    headers = ['Intern Name', 'Email', 'Days Present', 'Days Late', 'Days Absent', 'Logs Submitted', 'Tasks Completed', 'Attendance %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Calculate date range
    start_date = datetime.fromisoformat(start_date_str)
    end_date = datetime.fromisoformat(end_date_str)
    delta = (end_date - start_date).days + 1
    
    row = 4
    for intern in interns:
        present_count = 0
        late_count = 0
        absent_count = 0
        logs_count = 0
        tasks_count = 0
        
        # Loop through each day in the week
        current_date = start_date
        for _ in range(delta):
            date_str = current_date.strftime('%Y-%m-%d')
            
            # Attendance
            attendance_records = attendance_repo.get_attendance_by_intern(intern['id'])
            attendance = next((a for a in attendance_records if a['date'] == date_str), None)
            if attendance:
                if attendance['status'] == 'present':
                    present_count += 1
                elif attendance['status'] == 'late':
                    late_count += 1
                else:
                    absent_count += 1
            else:
                absent_count += 1
            
            # Logs
            log = log_repo.get_log_by_intern_and_date(intern['id'], date_str)
            if log:
                logs_count += 1
            
            current_date += timedelta(days=1)
        
        # Tasks completed in the week
        tasks = task_repo.get_tasks_by_intern(intern['id'])
        for task in tasks:
            if task.get('completed_at'):
                completed_date = task['completed_at'][:10]
                if start_date_str <= completed_date <= end_date_str:
                    tasks_count += 1
        
        # Calculate attendance percentage
        attendance_pct = ((present_count + late_count) / delta * 100) if delta > 0 else 0
        
        ws.cell(row=row, column=1).value = intern['name']
        ws.cell(row=row, column=2).value = intern['email']
        ws.cell(row=row, column=3).value = present_count
        ws.cell(row=row, column=4).value = late_count
        ws.cell(row=row, column=5).value = absent_count
        ws.cell(row=row, column=6).value = logs_count
        ws.cell(row=row, column=7).value = tasks_count
        ws.cell(row=row, column=8).value = f"{attendance_pct:.1f}%"
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_monthly_report(month: str) -> BytesIO:
    """Generate monthly report"""
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws = wb.active
    ws.title = "Monthly Summary"
    
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    month_name = datetime.strptime(month + '-01', '%Y-%m-%d').strftime('%B %Y')
    ws['A1'] = f'MONTHLY REPORT - {month_name}'
    ws['A1'].font = Font(bold=True, size=16, color="0066CC")
    ws.merge_cells('A1:J1')
    
    # Get all interns
    interns = user_repo.get_users_by_role('intern')
    
    # Headers
    headers = ['Intern Name', 'Email', 'Attendance %', 'Logs Submitted', 'Tasks Completed', 'Projects', 'Expenses', 'Stipend', 'Leave Days', 'WFH Days']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    for intern in interns:
        # Attendance stats
        attendance_counts = attendance_repo.count_attendance_by_status(intern['id'], month)
        total_days = sum(attendance_counts.values())
        attendance_pct = ((attendance_counts['present'] + attendance_counts['late']) / total_days * 100) if total_days > 0 else 0
        
        # Logs
        logs = log_repo.get_logs_by_intern(intern['id'])
        month_logs = [l for l in logs if l['date'].startswith(month)]
        
        # Tasks
        tasks = task_repo.get_tasks_by_intern(intern['id'])
        completed_tasks = [t for t in tasks if t.get('completed_at', '').startswith(month)]
        
        # Projects
        projects = project_repo.get_projects_by_intern(intern['id'])
        month_projects = [p for p in projects if p['month'] == month]
        
        # Expenses
        expenses = expense_repo.get_expenses_by_intern(intern['id'])
        month_expenses = [e for e in expenses if e['date'].startswith(month) and e['status'] == 'approved']
        total_expenses = sum(float(e['amount']) for e in month_expenses)
        
        # Stipend
        stipend = stipend_repo.get_stipend(intern['id'], month)
        stipend_amount = stipend['total_amount'] if stipend else 0
        
        # Leave days
        leaves = leave_repo.get_leaves_by_intern(intern['id'])
        approved_leaves = [l for l in leaves if l['status'] == 'approved' and l['start_date'].startswith(month)]
        leave_days = sum(
            (datetime.fromisoformat(l['end_date']) - datetime.fromisoformat(l['start_date'])).days + 1
            for l in approved_leaves
        )
        
        # WFH days
        wfh_requests = wfh_repo.get_wfh_by_intern(intern['id'])
        wfh_days = len([w for w in wfh_requests if w['date'].startswith(month) and w['status'] == 'approved'])
        
        ws.cell(row=row, column=1).value = intern['name']
        ws.cell(row=row, column=2).value = intern['email']
        ws.cell(row=row, column=3).value = f"{attendance_pct:.1f}%"
        ws.cell(row=row, column=4).value = len(month_logs)
        ws.cell(row=row, column=5).value = len(completed_tasks)
        ws.cell(row=row, column=6).value = len(month_projects)
        ws.cell(row=row, column=7).value = f"₹{total_expenses:.2f}"
        ws.cell(row=row, column=8).value = f"₹{stipend_amount:.2f}"
        ws.cell(row=row, column=9).value = leave_days
        ws.cell(row=row, column=10).value = wfh_days
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output